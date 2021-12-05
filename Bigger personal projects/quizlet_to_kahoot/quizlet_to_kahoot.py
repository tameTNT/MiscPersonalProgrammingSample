# Last Edited: 15:04 BST - 03/07/2021
from typing import List

import openpyxl as xl
from random import choice, shuffle
import argparse


def get_n_random_items_excluding(options: List[tuple], excl: tuple, n: int, item_part: int,
                                 prefer_str: str = '', avoid_str: str = '', thoroughness: int = 100) -> tuple:
    """
    Picks n random items from options with no duplicates and attempting to favour any options with prefer_str within them.

    :param options: list of option pair tuples, i.e. [(KR, EN), (KR, EN), ...]
    :param excl: (KR, EN) pair to avoid completely - usually the 'correct answer'. Must be present in options
    :param n: number of items to choose
    :param item_part: index of items chosen to return, i.e. all the KRs (0) or all the ENs (1)
    :param prefer_str: the function will attempt to favour items that include prefer_str
    :param avoid_str: the function will attempt to avoid items that include avoid_str
    :param thoroughness: number of attempts function will make to try and find an item satisfying prefer and avoid strings
    :return: a tuple of the successfully chosen parts, e.g. (KR1, KR2, KR3) for n=3, item_part=0
    """
    reduced_options = list(options)
    reduced_options.remove(excl)
    reduced_options = [o[item_part] for o in reduced_options]

    item_set = set()
    backup_items = set()

    for _ in range(n):
        item_found = False

        rand_item = 'None'

        search_attempts = 0
        while not item_found and search_attempts < thoroughness:
            if reduced_options:
                rand_item = choice(reduced_options)
                item_found = True
            else:  # all options exhausted
                item_found = False
                break

            reduced_options.remove(rand_item)

            if (prefer_str and (prefer_str not in rand_item)) or (avoid_str and (avoid_str in rand_item)):
                backup_items.add(rand_item)  # item backed-up before removal
                item_found = False

            search_attempts += 1

        if not item_found:
            rand_item = backup_items.pop()

        item_set.add(rand_item)

    backup_items.update(set(reduced_options))  # merges all remaining options into backup set
    while len(item_set) != n:  # in case not enough matching items were found
        item_set.add(backup_items.pop())

    return tuple(item_set)


def shorten_str(s: str, max_len: int) -> str:
    """
    Truncates string, s, if it is longer than max_len, appending a '...' in this case.

    :param s: string to potentially shorten
    :param max_len: maximum length of returned string. 
        Even with appended '...', the length of the returned string will NEVER exceed max_len.
    :return: shortened string
    """

    return s[:max_len-3] + '...' if len(s) > max_len else s


def convert(quizlet_export_str: str, verbose: bool):
    kr_en_list: List[tuple] = [tuple(s.strip() for s in KR_EN.split('%t')) for KR_EN in quizlet_export_str.split('%r')]

    kr_def_dict = dict(kr_en_list)
    potential_items = list(kr_def_dict.items())
    shuffle(potential_items)  # shuffles in place

    num_questions = int(input('\nHow many Kahoot questions should be created? (selected randomly from list entered above) >'))
    if num_questions > len(potential_items):
        raise Exception("You can't have more questions than there are original term/definition pairs to pick from!")
    if num_questions > 100:
        raise Exception("Kahoot doesn't allow more than 100 questions. :(")

    translation_dir = int(input('\nWhich way should translations go? Please enter 1, 2 or 3.\n'
                                'Q->A [1: KR->EN, 2: EN->KR, 3: Both ways (randomly chosen)] >'))

    time_limit = int(input('\nPlease enter time limit (in seconds) to use for every question.\n'
                           '[5, 10, 20, 30, 60, 90, 120, 240] >'))

    # Grab the pre-downloaded xlsx Kahoot template
    wb = xl.open('KahootQuizTemplate.xlsx')
    ws = wb.active  # gets main worksheet from workbook

    for row_i, item in enumerate(potential_items[:num_questions]):
        if translation_dir == 3:  # 3: Both ways so choose which way round - asking to define KR or find KR matching given EN?
            question_type = choice(('KR', 'EN'))
        elif translation_dir == 1:  # 1 : KR->EN
            question_type = 'KR'
        else:  # 2: EN->KR
            question_type = 'EN'

        if question_type == 'KR':
            q_index, a_index = 0, 1
        else:  # question_type == 'EN'
            q_index, a_index = 1, 0

        # Questions go in column B (120 chars)
        # Questions start from row 9 onwards
        ws[f'B{row_i + 9}'] = shorten_str(f'Pick the {("KR", "EN")[a_index]} for "{item[q_index]}":', 120)

        # Choose alternative answers
        if question_type == 'KR' and item[q_index][-1] == '다':  # prefer getting EN verb answers (since KR question is likely a verb)
            alt_answers = get_n_random_items_excluding(potential_items, item, 3, a_index, prefer_str='to ')
        elif question_type == 'EN' and item[a_index][-1] == '다':  # prefer getting KR verbs answers (since actual answer is also likely a verb)
            alt_answers = get_n_random_items_excluding(potential_items, item, 3, a_index, prefer_str='다')
        else:  # a noun, pronoun, etc. - not a verb
            if question_type == 'KR':
                alt_answers = get_n_random_items_excluding(potential_items, item, 3, a_index, avoid_str='to ')
            else:  # question_type == 'EN':
                alt_answers = get_n_random_items_excluding(potential_items, item, 3, a_index, avoid_str='다')

        alt_answers = iter(alt_answers)  # makes alt_answers an iterable so next() can be called on it

        # Answers 1-4 go in column C-F (75 chars)
        option_columns = ('C', 'D', 'E', 'F')
        correct_answer_col = choice(option_columns)
        for column in option_columns:
            target_cell = f'{column}{row_i + 9}'
            if column == correct_answer_col:
                ws[target_cell] = shorten_str(item[a_index], 75)  # Set correct answer str

                # Correct answer number(s) in column H (comma separated)
                ws[f'H{row_i + 9}'] = option_columns.index(correct_answer_col) + 1  # mark correct answer number (+1 for zero-indexing)
            else:
                ws[target_cell] = shorten_str(next(alt_answers), 75)

        # Time limits go in column G (5, 10, 20, 30, 60, 90, 120, or 240)
        ws[f'G{row_i + 9}'] = time_limit  # Set time limit

        if verbose:
            print(f'Q{row_i+1}: {item[q_index]}\n{item[a_index]}')

    # Save the created file
    wb.save('completed_quizlet_to_kahoot_convert.xlsx')

    print(
        'Spreadsheet successfully created ("completed_quizlet_to_kahoot_convert.xlsx").\n'
        'Upload the .xlsx at https://create.kahoot.it/creator/'
    )


main_parser = argparse.ArgumentParser(description='')
main_parser.add_argument('-v', '--verbose', action='store_true',
                         help='If given, all options are printed to console as well as written the the .xlsx file')

args = main_parser.parse_args()

user_input = input(
    'Export Quizlet set with settings:\n'
    '  Between term (KR) and definition (EN): "%t"\n  Between rows: "%r"\n'
    'Paste string(s) >'
).strip('%r')  # removes trailing %r
convert(user_input, args.verbose)
